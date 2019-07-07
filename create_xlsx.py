import openpyxl as excel
import tkinter
from openfile import samplelist
from openfile import datatable_dict
from tkinter import filedialog

wb = excel.Workbook()
ws = wb.active

row_num = 2
y = row_num

column_num = 2
x = column_num

for sample in samplelist:
    ws.cell(y, x, sample)
    y += 1
    ws.cell(y, x, 'name')
    ws.cell(y, x+1, 'RT')
    ws.cell(y, x+2, 'RRT')
    ws.cell(y, x+3, 'Area')
    ws.cell(y, x+4, 'Area%')
    ws.cell(y, x+5, '補正Area%')
    # サンプル名をセルに記入し、一行下にname, RT, RRT, Area, Area%, 補正Area%を記入

    rt_list = [datalist[0] for datalist in datatable_dict[sample]]  # ２次元配列の各リストの１つ目の要素をリスト化（RT）
    area_list = [datalist[1] for datalist in datatable_dict[sample]]  # ２次元配列の各リストの2つ目の要素をリスト化（RT）
    areaper_list = [datalist[2] for datalist in datatable_dict[sample]]  # ２次元配列の各リストの3つ目の要素をリスト化（RT）

    x += 1  # １列右にずらしてnameの下からRTの下に移動
    for rt_data in rt_list:
        # rt_listからrt_dataを取り出し
        y += 1
        ws.cell(y, x, rt_data)
        # 行を下にずらしながらセルに記入

    y += 2
    ws.cell(y, x-1, '基準RT')

    std_rt_x = x
    std_rt_y = y
    # std_rtのセル位置をあとで使う用の変数

    y = 4  # 3行目に戻る(RT等のカラムタイトルが書かれている行の１個下)
    x += 1  # 1列右にずらしてRTの下からRRTの下に移動

    length = len(rt_list)
    while y < 4+length:
        ws.cell(y, x,
                '='
                + ws.cell(y, x-1).coordinate
                + '/'
                + ws.cell(std_rt_y, std_rt_x).coordinate)
        y += 1
        # RTをstd_RTで割ってRRTを計算させる

    y = 3
    x += 1  # 1列右にずらしてRRTの下からAreaの下に移動

    for area_data in area_list:
        # area_listからarea_dataを取り出し
        y += 1
        ws.cell(y, x, area_data)
        # 行を下にずらしながらセルに記入

    y += 2
    ws.cell(y, x,
            '=SUM('
            + ws.cell(4, x).coordinate
            + ':'
            + ws.cell(y-2, x).coordinate
            + ')')
    # SUM関数を文字列としてエクセルに記入,total areaを計算させる。

    ws.cell(y, x-1, 'total Area')
    totalarea_x = x
    totalarea_y = y
    # totalareaのセル位置をあとで使う用の変数

    y = 3
    x += 1  # 1列右にずらしてAreaの下からArea%の下に移動

    for areaper_data in areaper_list:
        # areaper_listからareaper_dataを取り出し
        y += 1
        ws.cell(y, x, areaper_data)
        # 行を下にずらしながらセルに記入

    y = 4
    x += 1  # 1列右にずらしてAreaの下からArea%の下に移動

    while y < 4+length:
        ws.cell(y, x,
                '='
                + ws.cell(y, x-2).coordinate
                + '/'
                + ws.cell(totalarea_y, totalarea_x).coordinate)
        y += 1

    x += 2
    y = 2
    # fordatalistから抜けたら列を右に2列ずらして、行を元（二行目）に戻す

root = tkinter.Tk()
root.savefile = filedialog.asksaveasfilename(initialdir="/", title="Save as", filetypes=[("xlsx file", "*.xlsx")])

wb.save(root.savefile)
