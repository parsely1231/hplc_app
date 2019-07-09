import os
import openpyxl as excel
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox


datatable = []
# 2次元配列用の空リスト

datatable_dict = {}
# 2次元配列を複数入れるための辞書

samplelist = []
# サンプルネームの空リスト

wb = excel.Workbook()
ws = wb.active

filepath = ''


def fileselect_button_clicked():  # 変換するファイルを選ぶ関数（ボタンを押した時に実行）
    ftype = [('text file', '*.txt')]
    idir = os.path.abspath(os.path.dirname(__file__))
    global filepath
    filepath = filedialog.askopenfilename(filetypes=ftype, initialdir=idir)
    filename.set(filepath)


def readfile():
    with open(filepath, 'r') as file_open:
        n = -1

        global datatable
        datatable = []

        global datatable_dict
        datatable_dict = {}

        global samplelist
        samplelist = []

        for line in file_open:
            if line[0] == '#':
                line = line.replace('\n', '')
                samplelist.append(line)
                n += 1
                continue

            elif line[0] in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '0']:
                str_list = line.split()
                float_list = [float(s) for s in str_list]
                datatable.append(float_list)
                # 読み込んだ一行のデータを、floatリストにする
                # その後２次元配列用の空リストに入れる
                continue

            elif not datatable == []:
                datatable_name = samplelist[n]
                datatable_dict[datatable_name] = datatable
                datatable = []
                # 直近のサンプル名をkeyにして作成中の２次元配列をvalueにする。
                # その後２次元配列の中身をからにする


def convert():
    row_num = 2
    y = row_num

    column_num = 2
    x = column_num

    for sample in samplelist:
        ws.cell(y, x, sample)
        y += 1
        ws.cell(y, x, 'name')
        ws.cell(y, x+1, 'RT')
        ws.cell(y, x+2, 'RRT')
        ws.cell(y, x+3, 'Area')
        ws.cell(y, x+4, 'Area%')
        ws.cell(y, x+5, '補正Area%')
        # サンプル名をセルに記入し、一行下にname, RT, RRT, Area, Area%, 補正Area%を記入

        rt_list = [datalist[0] for datalist in datatable_dict[sample]]  # ２次元配列の各リストの１つ目の要素をリスト化（RT）
        area_list = [datalist[1] for datalist in datatable_dict[sample]]  # ２次元配列の各リストの2つ目の要素をリスト化（RT）
        areaper_list = [datalist[2] for datalist in datatable_dict[sample]]  # ２次元配列の各リストの3つ目の要素をリスト化（RT）

        totalarea_formula = '=SUM(' + ws.cell(4, x+3).coordinate + ':' + ws.cell(3+len(area_list), x+3).coordinate + ')'

        for rt_data, area_data, areaper_data in zip(rt_list, area_list, areaper_list):
            # listからdataを取り出し
            y += 1
            rrt_formula = '=ROUND(' + ws.cell(y, x+1).coordinate + \
                          '/' + ws.cell(5 + len(rt_list), x+1).coordinate + ', 2)'  # formula(RT/stdRT)
            areaper_formula = '=' + ws.cell(y, x+3).coordinate + \
                              '/' + ws.cell(5+len(area_list), x+3).coordinate  # formula(Area/totalArea)

            ws.cell(y, x+1, rt_data)  # write RT
            ws.cell(y, x+2, rrt_formula)  # write RRT
            ws.cell(y, x+3, area_data)  # write Area
            ws.cell(y, x+4, areaper_data)  # write Area%
            ws.cell(y, x+5, areaper_formula)  # write 補正Area%
            # 行を下にずらしながらセルに記入

        ws.cell(y+2, x, '基準RT')
        ws.cell(y+2, x+2, 'total Area')
        ws.cell(y+2, x+3, totalarea_formula)
        # SUM関数を文字列としてエクセルに記入,total areaを計算させる。

        if rrt_check_val.get():  # rrt_checkboxがONの時、std_rtを記入（RTリストから近いのを選ぶ）
            fl_var = float(std_rt_entry.get())
            std_rt_list = [std_rt for std_rt in rt_list if fl_var - 0.2 < std_rt < fl_var + 0.2]

            if not std_rt_list == []:
                ws.cell(y+2, x+1, std_rt_list[0])
            else:
                pass

        x += 7
        y = 2
        # for xx_data から抜けたら列を右に6列ずらして、行を二行目に戻す


def convert_button_clicked():  # 選んだファイルを読み込んでエクセルファイルに変換する（ボタンを押した時に実行）
    readfile()
    convert()
    root.savefile = filedialog.asksaveasfilename(initialdir="/", title="Save as", filetypes=[("xlsx file", "*.xlsx")])
    wb.save(root.savefile)
    messagebox.showinfo('message', 'Operation has completed successfully')


if __name__ == '__main__':
    root = Tk()
    root.title(u'HPLC txt to xlsx')
    root.geometry('600x400')

    static1 = ttk.Label(root, text='変換元のテキストファイルを選んでください')
    static1.pack()

    filename = StringVar()
    filename_entry = ttk.Entry(root, textvariable=filename, width=40)
    filename_entry.pack()

    fileselect_button = ttk.Button(root, text=u'file select', width=10, command=fileselect_button_clicked)
    fileselect_button.pack()

    midframe = ttk.Frame(root, padding=20)
    midframe.pack()

    rrt_check_val = BooleanVar()
    rrt_check_val.set(False)
    rrt_check = ttk.Checkbutton(midframe, text=u'RRTを計算する', variable=rrt_check_val)
    rrt_check.grid(row=0, sticky=W)

    rrtlabel = ttk.Label(midframe, text='基準とするRTの値を入力して下さい')
    rrtlabel.grid(row=1, column=0, sticky=W)

    var = IntVar()
    std_rt_entry = ttk.Entry(midframe, textvariable=var, width=15)
    std_rt_entry.grid(row=1, column=1, sticky=W)

    infolabel = ttk.Label(midframe, text='※入力値±0.2から実測値を拾います')
    infolabel.grid(row=2, column=0, sticky=W)

    convert_button = ttk.Button(root, text=u'convert', width=10, command=convert_button_clicked)
    convert_button.pack()

    root.mainloop()
